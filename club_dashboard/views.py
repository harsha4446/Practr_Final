from django.shortcuts import render
from users.models import clubs, events, student,colleges, rounds, register_table,student, student_scores, event_registered, round_room, room_judge, round_scores, event_registered_details,sub_head, judge_detail, student_detail, follow_table, teams
from . forms import addEvent, addRound, rooms, deadlines, newJudge
from django.http import HttpResponseRedirect, HttpResponse
import time
from django.contrib import auth
from django.utils.encoding import smart_str
from wsgiref.util import FileWrapper
import mimetypes

# Create your views here.


def updatePractrScores(scores):
    for object in scores:
        studentscore,create = student_scores.objects.get_or_create(username=object.student)
        multiplier = rounds.objects.get(id=object.round.id)
        multiplier = multiplier.weight
        studentscore.creativity = int (studentscore.creativity +  object.creativity)
        studentscore.content = int (studentscore.content + object.content)
        studentscore.presentation = int (studentscore.presentation + object.presentation)
        studentscore.rebuttal = int (studentscore.rebuttal + object.rebuttal)
        studentscore.communication = int (studentscore.communication + object.communication)
        studentscore.feasibility = int (studentscore.feasibility + object.feasibility)
        studentscore.save()


def percentage(x,total):
    if x > 0:
        return int((float(x)/float(total))*100)
    else:
        return 0


def dashboard(request,core1='1',core2='2'):
    user = request.user
    if not user.club or user.judge:
        if user.college:
            return HttpResponseRedirect("/user/college_dashboard/")
        if user.judge:
            return HttpResponseRedirect("/user/judge_dashboard/")
        return HttpResponseRedirect("/user/student_dashboard/")
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    try:
        event = events.objects.get(email = club, current = True)
        count1 = rounds.objects.filter(email=event, type=1).count()
        count2 = rounds.objects.filter(email=event, type=2).count()
        count3 = rounds.objects.filter(email=event, type=3).count()
        count4 = rounds.objects.filter(email=event, type=4).count()
        count5 = rounds.objects.filter(email=event, type=5).count()
        count6 = rounds.objects.filter(email=event, type=6).count()
        count7 = rounds.objects.filter(email=event, type=7).count()
        count8 = rounds.objects.filter(email=event, type=8).count()
        count9 = rounds.objects.filter(email=event, type=9).count()
        registered = event_registered_details.objects.filter(event=event).order_by('-id')[:5]
        marketingcount = event_registered_details.objects.filter(event=event,marketing=True).count()
        financecount = event_registered_details.objects.filter(event=event,finance=True).count()
        prcount = event_registered_details.objects.filter(event=event,public_relations=True).count()
        hrcount = event_registered_details.objects.filter(event=event,human_resources=True).count()
        edcount = event_registered_details.objects.filter(event=event,ent_dev=True).count()
        bmcount = event_registered_details.objects.filter(event=event,best_manager=True).count()
        cscount = event_registered_details.objects.filter(event=event, corp_strg=True).count()
        qucount = event_registered_details.objects.filter(event=event, quiz=True).count()
        tecount = event_registered_details.objects.filter(event=event, team=True).count()
        total =  marketingcount+financecount+prcount+hrcount+edcount+bmcount+cscount+qucount+tecount
        marketingcount = percentage(marketingcount, total)
        financecount = percentage(financecount, total)
        prcount = percentage(prcount, total)
        hrcount = percentage(hrcount, total)
        edcount = percentage(edcount, total)
        bmcount = percentage(bmcount, total)
        cscount = percentage(cscount,total)
        qucount = percentage(qucount,total)
        tecount = percentage(tecount,total)
    except events.DoesNotExist:
        context= {"user":user, "club":club,}
        return render(request,'club_dash/noEvent.html',context)
    table1 = event_registered_details.objects.filter(event=event, marketing=True).order_by('-mkttotal')[:5]
    table2 = event_registered_details.objects.filter(event=event, finance=True).order_by('-fintotal')[:5]
    if core1 != '1' or core2 != '2':
        if core1 == '1':
            table1 = event_registered_details.objects.filter(event=event, marketing=True).order_by('-mkttotal')[:5]
        elif core1 == '2':
            table1 = event_registered_details.objects.filter(event=event, finance=True).order_by('-fintotal')[:5]
        elif core1 == '3':
            table1 = event_registered_details.objects.filter(event=event, public_relations=True).order_by('-prtotal')[:5]
        elif core1 == '4':
            table1 = event_registered_details.objects.filter(event=event, human_resources=True).order_by('-hrtotal')[:5]
        elif core1 == '5':
            table1 = event_registered_details.objects.filter(event=event, ent_dev=True).order_by('-edtotal')[:5]
        elif core1 == '6':
            table1 = event_registered_details.objects.filter(event=event, best_manager=True).order_by('-bmtotal')[:5]
        elif core1 == '7':
            table1 = event_registered_details.objects.filter(event=event, corp_strg=True).order_by('-cstotal')[:5]
        elif core1 == '8':
            table1 = event_registered_details.objects.filter(event=event, quiz=True).order_by('-qutotal')[:5]
        elif core1 == '9':
            table1 = event_registered_details.objects.filter(event=event, team=True).order_by('-tetotal')[:5]
        if core2 == '1':
            table2 = event_registered_details.objects.filter(event=event, marketing=True).order_by('-mkttotal')[:5]
        elif core2 == '2':
            table2 = event_registered_details.objects.filter(event=event, finance=True).order_by('-fintotal')[:5]
        elif core2 == '3':
            table2 = event_registered_details.objects.filter(event=event, public_relations=True).order_by('-prtotal')[:5]
        elif core2 == '4':
            table2 = event_registered_details.objects.filter(event=event, human_resources=True).order_by('-hrtotal')[:5]
        elif core2 == '5':
            table2 = event_registered_details.objects.filter(event=event, ent_dev=True).order_by('-edtotal')[:5]
        elif core2 == '6':
            table2 = event_registered_details.objects.filter(event=event, best_manager=True).order_by('-bmtotal')[:5]
        elif core2 == '7':
            table2 = event_registered_details.objects.filter(event=event, corp_strg=True).order_by('-cstotal')[:5]
        elif core2 == '8':
            table2 = event_registered_details.objects.filter(event=event, quiz=True).order_by('-qutotal')[:5]
        elif core2 == '9':
            table2 = event_registered_details.objects.filter(event=event, team=True).order_by('-tetotal')[:5]

    context = {"user":user, "event":event, "club":club,'count1':count1, 'count2':count2, 'count3':count3, 'count4':count4,
               'count5':count5, 'count6':count6,'count7':count7,'count8':count8,'count9':count9,'registered':registered,
               'table1':table1,'table2':table2,'core1':core1,'core2':core2,'marketingcount':marketingcount,
               'financecount':financecount,'prcount':prcount,'hrcount':hrcount,'edcount':edcount,'bmcount':bmcount,
               'cscount': cscount,'qucount':qucount,'tecount':tecount,}
    return render(request,'club_dash/dashboard.html',context)


def live(request, id):
    event = events.objects.get(id=id)
    if event.live:
        event.live = False
        event.current = False
        heads = sub_head.objects.filter(event=event)
        for student in heads:
            student.student.delete()
        registered = event_registered.objects.filter(registered_to=event)
        for obj in registered:
            details,created = event_registered_details.objects.get_or_create(student=obj.current_user, event=event)
            details.delete()
            obj.rmregister(obj.current_user,event)
            obj.save()
        #event.delete()
        event.save()
    else:
        event.live = True
        event.current = True
        event.save()
    return HttpResponseRedirect("/user/club_dashboard/")


def activate_registraion(request, id):
    event = events.objects.get(id=id)
    if event.registration:
        event.registration = False
    else:
        event.registration = True
    event.save()
    return HttpResponseRedirect("/user/club_dashboard/")


def add_event(request,access = None, id=None):
    club = clubs.objects.get(club_email=request.user.email)
    exiting = None
    if id != None:
        exiting = events.objects.get(id=id)
    form = addEvent(request.POST or None, request.FILES or None)
    if form.is_valid():
        if id != None:
            event = events.objects.get(id=id)
        else:
            event = events(email=club)
        event.name = form.cleaned_data.get("name")
        event.about = form.cleaned_data.get("about")
        event.website = form.cleaned_data.get("website")
        if access == '1':
            event.inter_type = True
        else:
            event.prefix = request.POST.get("prefix")
            event.prefix = event.prefix.upper()
            event.tprefix = request.POST.get("tprefix")
            event.tprefix = event.tprefix.upper()
            if request.POST.get("multi") != None:
                event.multiregistration = True
        event.team_size1 = request.POST.get("teamsize1")
        event.team_size2 = request.POST.get("teamsize2")
        event.team_size3 = request.POST.get("teamsize3")
        event.team_size4 = request.POST.get("teamsize4")
        event.team_size5 = request.POST.get("teamsize5")
        event.team_size6 = request.POST.get("teamsize6")
        event.team_size7 = request.POST.get("teamsize7")
        event.team_size8 = request.POST.get("teamsize8")
        event.team_size9 = request.POST.get("teamsize9")
        if request.POST.get("mktlabel") != '':
            event.mktlabel = request.POST.get("mktlabel", "Marketing")
        if request.POST.get("finlabel") != '':
            event.finlabel = request.POST.get("finlabel", "Finance")
        if request.POST.get("prlabel") != '':
            event.prlabel = request.POST.get("prlabel", "Public Relation")
        if request.POST.get("hrlabel") != '':
            event.hrlabel = request.POST.get("hrlabel", "Human Resource")
        if request.POST.get("edlabel") != '':
            event.edlabel = request.POST.get("edlabel", "Entrepreneurship Development")
        if request.POST.get("bmlabel") != '':
            event.bmlabel = request.POST.get("bmlabel", "Best Manager")
        if request.POST.get("cslabel") != '':
            event.cslabel = request.POST.get("cslabel", "Corporate Strategy")
        if request.POST.get("qulabel") != '':
            event.qulabel = request.POST.get("qulabel", "Quiz")
        if request.POST.get("telabel") != '':
            event.telabel = request.POST.get("telabel", "Team")
        event.marketing = form.cleaned_data.get("marketing")
        event.finance = form.cleaned_data.get("finance")
        event.public_relations = form.cleaned_data.get("public_relations")
        event.human_resources = form.cleaned_data.get("human_resources")
        event.ent_dev = form.cleaned_data.get("ent_dev")
        event.best_manager = form.cleaned_data.get("best_manager")
        event.corp_strg = form.cleaned_data.get("corp_strg")
        event.quiz = form.cleaned_data.get("quiz")
        event.team = form.cleaned_data.get("team")
        event.firstyear = request.POST.get("first",False)
        event.secondyear = request.POST.get("second",False)
        event.thirdyear = request.POST.get("third",False)
        if not event.firstyear:
            event.quota11 = 0
            event.quota21 = 0
            event.quota31 = 0
            event.quota41 = 0
            event.quota51 = 0
            event.quota61 = 0
            event.quota71 = 0
            event.quota81 = 0
            event.quota91 = 0
        if not event.secondyear:
            event.quota12 = 0
            event.quota22 = 0
            event.quota32 = 0
            event.quota42 = 0
            event.quota52 = 0
            event.quota62 = 0
            event.quota72 = 0
            event.quota82 = 0
            event.quota92 = 0
        if not event.thirdyear:
            event.quota13 = 0
            event.quota23 = 0
            event.quota33 = 0
            event.quota43 = 0
            event.quota53 = 0
            event.quota63 = 0
            event.quota73 = 0
            event.quota83 = 0
            event.quota93 = 0
        if form.cleaned_data.get("logo"):
            event.logo = form.cleaned_data.get("logo")
        event.current = True
        event.save()
        return HttpResponseRedirect("/user/club_dashboard")
    context = {"form":form, "user":request.user, 'existing':exiting,'access':access}
    return render(request, 'club_dash/add_inter_event.html', context)


def add_round(request, id=None, operation=None, offline=None,existing=None):
    if request.user.judge:
        details = judge_detail.objects.get(email=request.user)
        club = clubs.objects.get(id=details.club.id)
    else:
        club = clubs.objects.get(club_email = request.user.email)
    event = events.objects.get(id=id)
    if existing == None:
        round = rounds(email = event)
    else:
        round = rounds.objects.get(id=existing)
        existing = round
    form = addRound(request.POST or None)
    if form.is_valid():
        round.title = form.cleaned_data.get("title","")
        round.sub_title = form.cleaned_data.get("sub_title","")
        if request.POST.get("low"):
            round.weight = 0.33
        if request.POST.get("medium"):
            round.weight = 0.66
        if request.POST.get("high"):
            round.weight = 1.0
        round.club = club
        round.about = form.cleaned_data.get("about","")
        round.task1 = form.cleaned_data.get("task1","")
        round.task2 = form.cleaned_data.get("task2","")
        round.task3 = form.cleaned_data.get("task3","")
        round.task4 = form.cleaned_data.get("task4","")
        round.task5 = form.cleaned_data.get("task5","")
        round.resource1 = form.cleaned_data.get("resource1","")
        round.resource2 = form.cleaned_data.get("resource2","")
        round.resource3 = form.cleaned_data.get("resource3","")
        round.resource4 = form.cleaned_data.get("resource4","")
        round.resource5 = form.cleaned_data.get("resource5","")
        round.creativity = form.cleaned_data.get("creativity","")
        round.content = form.cleaned_data.get("content")
        round.presentation = form.cleaned_data.get("presentation")
        round.rebuttal = form.cleaned_data.get("rebuttal")
        round.communication = form.cleaned_data.get("communication")
        round.feasibility = form.cleaned_data.get("feasibility")
        round.feedback = form.cleaned_data.get("feedback")
        round.question1 = form.cleaned_data.get("question1","")
        round.question2 = form.cleaned_data.get("question2","")
        round.question3 = form.cleaned_data.get("question3","")
        round.question4 = form.cleaned_data.get("question4","")
        round.question5 = form.cleaned_data.get("question5","")
        round.core1 = form.cleaned_data.get("core1",0)
        round.core2 = form.cleaned_data.get("core2",0)
        round.core3 = form.cleaned_data.get("core3",0)
        round.core4 = form.cleaned_data.get("core4",0)
        round.core5 = form.cleaned_data.get("core5",0)
        round.core6 = form.cleaned_data.get("core6",0)
        round.creativityvalue = request.POST.get("creativityvalue",0)
        round.contentvalue = request.POST.get("contentvalue",0)
        round.presentationvalue = request.POST.get("presentationvalue",0)
        round.rebuttalvalue = request.POST.get("rebuttalvalue",0)
        round.communicationvalue = request.POST.get("communicationvalue",0)
        round.feasibilityvalue = request.POST.get("feasibilityvalue",0)
        round.weight = request.POST.get("weight")
        round.type = operation
        round.author = request.user.name
        round.offline = offline
        round.max_score = request.POST.get("max_score",100)
        if offline == '2':
            round.question1 = "Score Obtained"
            round.core1 = round.max_score
        round.save()
        type = int (operation)
        audience = round_scores.objects.filter(round__email=event, round__type=type, qualified=True)
        for student in audience:
            code = event_registered_details.objects.get(student=student.student,event=event)
            x,created = round_scores.objects.get_or_create(student=student.student, round=round, rcode=code.rcode)
            if offline != '0':
                x.submitted = True
            x.save()
        return HttpResponseRedirect("/user/club_dashboard/caseView/"+id+"/"+operation)
    context = {"form":form, "offline":offline,'existing':existing}
    return render (request,'club_dash/add_round.html',context)


def del_round(request, id = None):
    round = rounds.objects.get(id=id)
    event = str (round.email.id)
    type = str (round.type)
    round.delete()
    return HttpResponseRedirect("/user/club_dashboard/caseView/"+event+"/"+type)


def case_view(request, id, type):
    event = events.objects.get(id=id)
    all_rounds = rounds.objects.filter(email=event,type=type)
    roomForm = rooms(request.POST or None)
    deadlineForm = deadlines(request.POST or None)
    judgeForm = newJudge(request.POST or None)
    register = 0
    all_rooms = None
    corename=''
    try:
        if type == '1':
            register = event_registered_details.objects.filter(event=event,marketing=True).count()
            corename = event.mktlabel
        elif type == '2':
            register = event_registered_details.objects.filter(event=event, finance=True).count()
            corename = event.finlabel
        elif type == '3':
            register = event_registered_details.objects.filter(event=event, public_relations=True).count()
            corename = event.prlabel
        elif type == '4':
            register = event_registered_details.objects.filter(event=event, human_resources=True).count()
            corename = event.hrlabel
        elif type == '5':
            register = event_registered_details.objects.filter(event=event, ent_dev=True).count()
            corename = event.edlabel
        elif type == '6':
            register = event_registered_details.objects.filter(event=event, best_manager=True).count()
            corename = event.bmlabel
        elif type == '7':
            register = event_registered_details.objects.filter(event=event, corp_strg=True).count()
            corename = event.cslabel
        elif type == '8':
            register = event_registered_details.objects.filter(event=event, quiz=True).count()
            corename = event.qulabel
        elif type == '9':
            register = event_registered_details.objects.filter(event=event, team=True).count()
            corename = event.telabel
        all_rooms = round_room.objects.all()
    except register_table.DoesNotExist:
        register = 0
    context = {'all_rounds':all_rounds, 'user':request.user, 'event':event, 'type':type, 'count':register, 'roomForm':roomForm,
               'deadline':deadlineForm, 'all_rooms':all_rooms, 'judgeForm':judgeForm,'corename':corename}
    return render(request,'club_dash/case_view.html',context)


def publish(request, id=None, event=None, type=None):
    round = rounds.objects.get(id=id)
    event_detail = events.objects.get(id=event)
    if round.published:
        round.published = False
        round.deadline = None
    else:
        if request.POST:
            round.published = True
            round.deadline = request.POST['deadline']
            if type == '1':
                audience = event_registered_details.objects.filter(event=event,marketing=True)
            elif type == '2':
                audience = event_registered_details.objects.filter(event=event,finance=True)
            elif type == '3':
                audience = event_registered_details.objects.filter(event=event,public_relations=True)
            elif type == '4':
                audience = event_registered_details.objects.filter(event=event,human_resources=True)
            elif type == '5':
                audience = event_registered_details.objects.filter(event=event,ent_dev=True)
            elif type == '6':
                audience = event_registered_details.objects.filter(event=event,best_manager=True)
            elif type == '7':
                audience = event_registered_details.objects.filter(event=event,corp_strg=True)
            elif type == '8':
                audience = event_registered_details.objects.filter(event=event,quiz=True)
            else:
                audience = event_registered_details.objects.filter(event=event, team=True)
            for student in audience:
                x, created = round_scores.objects.get_or_create(student=student.student, round=round, rcode=student.rcode)
                if round.offline != '0':
                    x.submitted = True
                x.save()
            if round.offline != '0':
                for student in audience:
                    score,created = round_scores.objects.get_or_create(student=student.student,round=round)
                    score.submitted = True
                    score.rcode = student.rcode
                    score.save()
    round.save()
    return HttpResponseRedirect('/user/club_dashboard/caseView/'+event+'/'+type)


def members(request, id, type):
    current_user = request.user
    # round = rounds.objects.get(id=id)
    event = events.objects.get(id=id)
    audience = None
    # if round.finished:
    #     audience = round_scores.objects.filter(round=round,submitted=True)
    # else:
    if type == '1':
        audience = event_registered_details.objects.filter(event=event,marketing=True)
    elif type == '2':
        audience = event_registered_details.objects.filter(event=event,finance=True)
    elif type == '3':
        audience = event_registered_details.objects.filter(event=event,public_relations=True)
    elif type == '4':
        audience = event_registered_details.objects.filter(event=event,human_resources=True)
    elif type == '5':
        audience = event_registered_details.objects.filter(event=event,ent_dev=True)
    elif type == '6':
        audience = event_registered_details.objects.filter(event=event,best_manager=True)
    elif type == '7':
        audience = event_registered_details.objects.filter(event=event,corp_strg=True)
    elif type == '8':
        audience = event_registered_details.objects.filter(event=event,quiz=True)
    elif type == '9':
        audience = event_registered_details.objects.filter(event=event,team=True)
    context = {'user':current_user, 'audience':audience,'event':event}
    return render(request,'club_dash/members.html',context)


def result(request, round=None):
    user = request.user
    round = rounds.objects.get(id=round)
    event = round.email
    type = round.type
    scores = round_scores.objects.filter(round=round).order_by('-total')
    count = round_scores.objects.filter(round=round).count()
    context = {'user': user, 'round':round, 'scores':scores,'event':event,'type':type,'max':count}
    return render(request, 'club_dash/results.html', context)


def addRoom(request,id,event,type):
    round = rounds.objects.get(id = id)
    room = round_room(round = round)
    room.room = request.POST['room']
    room.save()
    return HttpResponseRedirect('/user/club_dashboard/caseView/'+event+'/'+type)


def addJudge(request,id,event,type):
    round = rounds.objects.get(id=id)
    room = round_room.objects.first()
    judge = room_judge(room=room, round=round)
    judge.judge_email = request.POST['judge_email']
    judge.judge_password = request.POST['judge_password']
    judge.save()
    user = student.objects.create_user(email=judge.judge_email, password=judge.judge_password, name='', phoneno='')
    user.judge = True
    user.save()
    return HttpResponseRedirect('/user/club_dashboard/caseView/' + event + '/' + type)


def teamSize(request, id, event,type):
    round = rounds.objects.get(id=id)
    round.team_size = request.POST.get('team_size')
    round.save()
    return HttpResponseRedirect('/user/club_dashboard/caseView/'+event+'/'+type)


def finishRound(request, id=None):
    round = rounds.objects.get(id=id)
    round.finished = True
    parameter = round_scores.objects.filter(round=round)
    updatePractrScores(parameter)
    round.save()
    scores = round_scores.objects.filter(round=round)
    for score in scores:
        student = event_registered_details.objects.get(student=score.student, event=round.email)
        score.total = float(float(round.weight) * score.total)
        score.save()
        if score.round.type == 1:
            student.mkttotal = student.mkttotal + score.total
        if score.round.type == 2:
            student.fintotal = student.fintotal + score.total
        if score.round.type == 3:
            student.prtotal = student.prtotal + score.total
        if score.round.type == 4:
            student.hrtotal = student.hrtotal + score.total
        if score.round.type == 5:
            student.edtotal = student.edtotal + score.total
        if score.round.type == 6:
            student.bmtotal = student.bmtotal + score.total
        if score.round.type == 7:
            student.cstotal = student.cstotal + score.total
        if score.round.type == 8:
            student.qutotal = student.qutotal + score.total
        if score.round.type == 9:
            student.tetotal = student.tetotal + score.total
        student.save()
    return HttpResponseRedirect('/user/club_dashboard/caseView/' + str(round.email.id) + '/' + str(round.type))


def audience(request, event, type):
    qualify = request.POST.get("qualified")
    mainevent = events.objects.get(id=event)
    qualify = int (qualify)
    qualified = None
    disqualified = None
    if type == '1':
        qualified = event_registered_details.objects.filter(event=mainevent,marketing=True).order_by('-mkttotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent,marketing=True)
    if type == '2':
        qualified = event_registered_details.objects.filter(event=mainevent, finance=True).order_by('-fintotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, finance=True)
    if type == '3':
        qualified = event_registered_details.objects.filter(event=mainevent, public_relations=True).order_by('-prtotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, public_relations=True)
    if type == '4':
        qualified = event_registered_details.objects.filter(event=mainevent, human_resources=True).order_by('-hrtotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, human_resources=True)
    if type == '5':
        qualified = event_registered_details.objects.filter(event=mainevent, ent_dev=True).order_by('-edtotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, ent_dev=True)
    if type == '6':
        qualified = event_registered_details.objects.filter(event=mainevent, best_manager=True).order_by('-bmtotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, best_manager=True)
    if type == '7':
        qualified = event_registered_details.objects.filter(event=mainevent, corp_strg=True).order_by('-cstotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, corp_strg=True)
    if type == '8':
        qualified = event_registered_details.objects.filter(event=mainevent, quiz=True).order_by('-qutotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, quiz=True)
    if type == '9':
        qualified = event_registered_details.objects.filter(event=mainevent, team=True).order_by('-tetotal')[
                    :qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, team=True)
    for student in disqualified:
        if student not in qualified:
            if type == '1':
                student.marketing = False
            if type == '2':
                student.finance = False
            if type == '3':
                student.public_relations = False
            if type == '4':
                student.human_resources = False
            if type == '5':
                student.ent_dev = False
            if type == '6':
                student.best_manager = False
            if type == '7':
                student.corp_strg = False
            if type == '8':
                student.quiz = False
            if type == '9':
                student.team = False
            student.save()
            scores = round_scores.objects.filter(round__email=mainevent, student=student.student, round__type=int(type))
            for obj in scores:
                obj.qualified = False
                obj.save()
            #final = event_registered.objects.get(current_user=student.student)
            #event_registered.rmregister(final.current_user,mainevent)
    return HttpResponseRedirect('/user/club_dashboard/masterTable/' + type)


def interAudience(request, event, type):
    qualify = request.POST.get("qualified")
    mainevent = events.objects.get(id=event)
    qualify = int (qualify)
    qualified = None
    disqualified = None
    rcodes = None
    if type == '1':
        qualified = event_registered_details.objects.filter(event=mainevent, marketing=True).values('rcode','mkttotal').distinct()
        qualified = qualified.order_by('-mkttotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent,marketing=True)
    if type == '2':
        qualified = event_registered_details.objects.filter(event=mainevent, finance=True).values('rcode','fintotal').distinct()
        qualified = qualified.order_by('-fintotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, finance=True)
    if type == '3':
        qualified = event_registered_details.objects.filter(event=mainevent, public_relations=True).values('rcode','prtotal').distinct()
        qualified = qualified.order_by('-prtotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, public_relations=True)
    if type == '4':
        qualified = event_registered_details.objects.filter(event=mainevent, human_resources=True).values('rcode','hrtotal').distinct()
        qualified = qualified.order_by('-hrtotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, human_resources=True)
    if type == '5':
        qualified = event_registered_details.objects.filter(event=mainevent, ent_dev=True).values('rcode','edtotal').distinct()
        qualified = qualified.order_by('-edtotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, ent_dev=True)
    if type == '6':
        qualified = event_registered_details.objects.filter(event=mainevent, best_manager=True).values('rcode','bmtotal').distinct()
        qualified = qualified.order_by('-bmtotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, best_manager=True)
    if type == '7':
        qualified = event_registered_details.objects.filter(event=mainevent, corp_strg=True).values('rcode','cstotal').distinct()
        qualified = qualified.order_by('-cstotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, corp_strg=True)
    if type == '8':
        qualified = event_registered_details.objects.filter(event=mainevent, quiz=True).values('rcode','qutotal').distinct()
        qualified = qualified.order_by('-qutotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, quiz=True)
    if type == '9':
        qualified = event_registered_details.objects.filter(event=mainevent, team=True).values('rcode','tetotal').distinct()
        qualified = qualified.order_by('-tetotal')[:qualify]
        disqualified = event_registered_details.objects.filter(event=mainevent, team=True)
    for student in disqualified:
        flag = True
        for x in qualified:
            if student.rcode == x['rcode']:
                flag = False
        if flag:
            if type == '1':
                student.marketing = False
            if type == '2':
                student.finance = False
            if type == '3':
                student.public_relations = False
            if type == '4':
                student.human_resources = False
            if type == '5':
                student.ent_dev = False
            if type == '6':
                student.best_manager = False
            if type == '7':
                student.corp_strg = False
            if type == '8':
                student.quiz = False
            if type == '9':
                student.team = False
            student.save()
            scores = round_scores.objects.filter(round__email=mainevent, student=student.student, round__type=int(type))
            for obj in scores:
                obj.qualified = False
                obj.save()
            #final = event_registered.objects.get(current_user=student.student)
            #event_registered.rmregister(final.current_user,mainevent)
    return HttpResponseRedirect('/user/club_dashboard/masterTable/' + type)


def eliminate(request, event, type, rcode):
    mainevent = events.objects.get(id=event)
    registered = event_registered_details.objects.filter(rcode=rcode,event=mainevent)
    for student in registered:
        if type == '1':
            student.marketing = False
        if type == '2':
            student.finance = False
        if type == '3':
            student.public_relations = False
        if type == '4':
            student.human_resources = False
        if type == '5':
            student.ent_dev = False
        if type == '6':
            student.best_manager = False
        if type == '7':
            student.corp_strg = False
        if type == '8':
            student.quiz = False
        if type == '9':
            student.team = False
        student.save()
        scores = round_scores.objects.filter(round__email=mainevent, student=student.student, round__type=int(type))
        for obj in scores:
            obj.qualified = False
            obj.save()
    return HttpResponseRedirect('/user/club_dashboard/masterTable/' + type)


def addSub(request,id,type):
    event = events.objects.get(id=id)
    username = request.POST.get("email")
    password = request.POST.get("password")
    name = request.POST.get("name")
    phone = request.POST.get("phone")
    head = student.objects.create_user(email=username,password=password,phoneno=phone,name=name)
    head.club = True
    head.judge = True
    head.activated = True
    head.save()
    set_type = judge_detail(email=head)
    set_type.type = type
    club = clubs.objects.get(club_email=request.user.email)
    set_type.club = club
    set_type.save()
    details = sub_head(student=head,type=type, event=event)
    details.save()
    if type == '1':
        event.subhead1 = True
    if type == '2':
        event.subhead2 = True
    if type == '3':
        event.subhead3 = True
    if type == '4':
        event.subhead4 = True
    if type == '5':
        event.subhead5 = True
    if type == '6':
        event.subhead6 = True
    if type == '7':
        event.subhead7 = True
    if type == '8':
        event.subhead8 = True
    if type == '9':
        event.subhead9 = True
    event.save()
    judge_setup = judge_detail(email=head,club=club,type=type)
    judge_setup.save()
    return HttpResponseRedirect("/user/club_dashboard/")


def quotaset(request,id=None):
    event = events.objects.get(id=id)
    event.quota11 = request.POST.get("quota11",0)
    event.quota21 = request.POST.get("quota21",0)
    event.quota31 = request.POST.get("quota31",0)
    event.quota41 = request.POST.get("quota41",0)
    event.quota51 = request.POST.get("quota51",0)
    event.quota61 = request.POST.get("quota61",0)
    event.quota71 = request.POST.get("quota71", 0)
    event.quota81 = request.POST.get("quota81", 0)
    event.quota91 = request.POST.get("quota91", 0)
    event.quota12 = request.POST.get("quota12",0)
    event.quota22 = request.POST.get("quota22",0)
    event.quota32 = request.POST.get("quota32",0)
    event.quota42 = request.POST.get("quota42",0)
    event.quota52 = request.POST.get("quota52",0)
    event.quota62 = request.POST.get("quota62",0)
    event.quota72 = request.POST.get("quota72", 0)
    event.quota82 = request.POST.get("quota82", 0)
    event.quota92 = request.POST.get("quota92", 0)
    event.quota13 = request.POST.get("quota13",0)
    event.quota23 = request.POST.get("quota23",0)
    event.quota33 = request.POST.get("quota33",0)
    event.quota43 = request.POST.get("quota43",0)
    event.quota53 = request.POST.get("quota53",0)
    event.quota63 = request.POST.get("quota63",0)
    event.quota73 = request.POST.get("quota73", 0)
    event.quota83 = request.POST.get("quota83", 0)
    event.quota93 = request.POST.get("quota93", 0)
    event.prefix = request.POST.get("prefix",'')
    event.save()
    time.sleep(2)
    return HttpResponseRedirect("/user/club_dashboard/")


def registered_members(request,id):
    user = request.user
    event = events.objects.get(id=id)
    registered = event_registered_details.objects.filter(event=event).order_by('-id')
    context = {'user':user,'registered':registered,'event':event}
    return render(request, 'club_dash/registeredmembers.html',context)


def master_table(request, type):
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    event = events.objects.get(email=club, current=True)
    type = int(type)
    teamsize = 1
    if type == 1:
        teamsize = event.team_size1
    elif type == 2:
        teamsize = event.team_size2
    elif type == 3:
        teamsize = event.team_size3
    elif type == 4:
        teamsize = event.team_size4
    elif type == 5:
        teamsize = event.team_size5
    elif type == 6:
        teamsize = event.team_size6
    elif type == 7:
        teamsize = event.team_size7
    elif type == 8:
        teamsize = event.team_size8
    elif type == 9:
        teamsize = event.team_size9
    if event.inter_type or teamsize != 1:
        registered = round_scores.objects.filter(round__email=event, round__type=type, qualified=True)
        distinct_registered = round_scores.objects.filter(round__email=event, round__type=type, qualified=True).values(
            'rcode','total','round').distinct()
        distinct_codes = distinct_registered.values('rcode').distinct()
        if type == 1:
            score_total = event_registered_details.objects.filter(event=event,regmarketing=True).values('rcode','mkttotal').distinct()
        elif type == 2:
            score_total = event_registered_details.objects.filter(event=event,regmarketing=True).values('rcode','fintotal').distinct()
        elif type == 3:
            score_total = event_registered_details.objects.filter(event=event,regpublic_relations=True).values('rcode','prtotal').distinct()
        elif type == 4:
            score_total = event_registered_details.objects.filter(event=event,reghuman_resources=True).values('rcode','hrtotal').distinct()
        elif type == 5:
            score_total = event_registered_details.objects.filter(event=event,regent_dev=True).values('rcode','edtotal').distinct()
        elif type == 6:
            score_total = event_registered_details.objects.filter(event=event,regbest_manager=True).values('rcode','bmtotal').distinct()
        elif type == 7:
            score_total = event_registered_details.objects.filter(event=event,regcorp_strg=True).values('rcode','cstotal').distinct()
        elif type == 8:
            score_total = event_registered_details.objects.filter(event=event,regquiz=True).values('rcode','qutotal').distinct()
        else:
            score_total = event_registered_details.objects.filter(event=event,regteam=True).values('rcode','tetotal').distinct()
        all_rounds = rounds.objects.filter(email=event, type=type, finished=True).order_by('-created')
        total = distinct_codes.count()
        dis_registered = round_scores.objects.filter(round__email=event, round__type=type, qualified=False)
        dis_distinct_registered = round_scores.objects.filter(round__email=event, round__type=type,
                                                              qualified=False).values(
            'rcode','total','round').distinct()
        dis_distinct_codes = dis_distinct_registered.values('rcode').distinct()
        context = {'user': user, 'registered': registered, 'max': total, 'event': event, 'type': type,
                   'rounds': all_rounds, 'dist_codes':distinct_codes,'dis_dist_codes':dis_distinct_codes,
                   'distinct': distinct_registered, 'scoretotal': score_total, 'dis_registered': dis_registered,
                   'dis_distinct': dis_distinct_registered}
        return render(request, 'club_dash/inter_audience_master.html', context)
    else:
        registered = round_scores.objects.filter(round__email=event, round__type=type,qualified=True)
        distinct_registered = round_scores.objects.filter(round__email=event, round__type=type,qualified=True).values('rcode','student__name','student__phoneno','student__email').distinct()
        score_total = event_registered_details.objects.filter(event=event)
        all_rounds = rounds.objects.filter(email=event, type=type,finished=True).order_by('-created')
        total = distinct_registered.count()
        #cases = rounds.objects.filter(email=event,type=type,published=True,finished=False)
        dis_registered = round_scores.objects.filter(round__email=event, round__type=type, qualified=False)
        dis_distinct_registered = round_scores.objects.filter(round__email=event, round__type=type, qualified=False).values(
            'rcode', 'student__name', 'student__phoneno', 'student__email').distinct()
        context ={'user':user,'registered':registered,'max':total,'event':event,'type':type,'rounds':all_rounds,
                    'distinct':distinct_registered,'scoretotal':score_total,'dis_registered':dis_registered,
                    'dis_distinct':dis_distinct_registered}
    return render(request,'club_dash/audience_master.html',context)


def edit_profile(request):
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    flag = 0
    try:
        event = events.objects.get(email=club,current=True)
    except events.DoesNotExist:
        event = None
    if request.POST:
        club.name = request.POST.get("club_name")
        club.website = request.POST.get("website",'')
        club.phone = request.POST.get("phone")
        club.video = request.POST.get("video", '')
        club.about = request.POST.get("about")
        club.admin_name = request.POST.get("admin_name")
        club.save()
        user.phoneno = club.phone
        if request.POST.get("old_password") != '' and request.POST.get("new_password") != '':
            old = request.POST.get("old_password")
            new = request.POST.get("new_password")
            if auth.authenticate(email=user.email, password=old):
                user.set_password(new)
                flag = 1
                user.save()
                auth.login(request, user)
            else:
                flag = 2
        user.name = club.admin_name
        user.save()
    context = {'user':user,'club':club,'flag':flag,'event':event}
    return render(request,'club_dash/edit_profile.html',context)


def download(request,id=None,file=0):
    scores = round_scores.objects.get(id=id)
    if file == '1':
        path = scores.data1.path
        name = scores.data1.name
        size = scores.data1.size
    elif file == '2':
        path = scores.data2.path
        name = scores.data2.name
        size = scores.data2.size
    else:
        path = scores.data3.path
        name = scores.data3.name
        size = scores.data3.size
    wrapper = FileWrapper(open(path, "rb" ))
    response = HttpResponse(wrapper,content_type=mimetypes.guess_type(path))
    #response['Content-Type'] = "application/force-download"
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(name)
    response['Content-Length'] = size
    return response


def export_scores(request,type=None):
    import openpyxl
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Scores.xlsx'
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    event = events.objects.get(email=club,current=True)
    all_rounds = rounds.objects.filter(email=event,type=type,finished=True).order_by('-created')
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Scores"
    row_num = 0
    rcode = ""
    columns = [
        (u"ID", 80),
        (u"Name", 20),
        (u"Email", 20),
        (u"Phone", 20),
    ]
    for round in all_rounds:
        columns = columns + [(u""+round.title,20),]
    columns = columns + [(u"Total",20),]
    for col_num in range(0,len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
    if type == '1':
        queryset = event_registered_details.objects.filter(marketing=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=1)
    elif type == '2':
        queryset = event_registered_details.objects.filter(finance=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=2)
    elif type == '3':
        queryset = event_registered_details.objects.filter(public_relations=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=3)
    elif type == '4':
        queryset = event_registered_details.objects.filter(human_resources=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=4)
    elif type == '5':
        queryset = event_registered_details.objects.filter(ent_dev=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=5)
    elif type == '6':
        queryset = event_registered_details.objects.filter(best_manager=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=6)
    elif type == '7':
        queryset = event_registered_details.objects.filter(corp_strg=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=7)
    elif type == '8':
        queryset = event_registered_details.objects.filter(quiz=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=8)
    else:
        queryset = event_registered_details.objects.filter(team=True,event=event).order_by('rcode')
        scores = round_scores.objects.filter(round__type=9)
    for obj in queryset:
        if rcode != obj.rcode:
            rcode = obj.rcode
            row_num += 1
            row = [
                obj.rcode,
                obj.student.name,
                obj.student.email,
                obj.student.phoneno,
            ]
            for x in all_rounds:
                for score in scores:
                    if score.round == x:
                        if score.student == obj.student:
                            row = row + [score.total,]
            if type == '1':
                row = row + [obj.mkttotal,]
            elif type == '2':
                row = row + [obj.fintotal,]
            elif type == '3':
                row = row + [obj.prtotal,]
            elif type == '4':
                row = row + [obj.hrtotal,]
            elif type == '5':
                row = row + [obj.edtotal,]
            elif type == '6':
                row = row + [obj.bmtotal,]
            elif type == '7':
                row = row + [obj.cstotal,]
            elif type == '8':
                row = row + [obj.qutotal,]
            else:
                row = row + [obj.tetotal,]
            for col_num in range(0,len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
    wb.save(response)
    return response


def export_registrations(request):
    import openpyxl
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Registrations.xlsx'
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    event = events.objects.get(email=club, current=True)
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Registrations"
    row_num = 0
    if event.inter_type:
        columns = [
            (u"ID", 80),
            (u"Name", 20),
            (u"Email", 20),
            (u"Phone", 20),
            (u"Course", 20),
            (u"Year", 20),
            (u"College", 20),
            (u"Core Events", 20),
        ]
    else:
        columns = [
            (u"ID", 80),
            (u"Name", 20),
            (u"Email", 20),
            (u"Phone", 20),
            (u"Course", 20),
            (u"Year", 20),
            (u"Section", 20),
            (u"Core Events", 20),
        ]
    for col_num in range(0,len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
    queryset = event_registered_details.objects.filter(event=event).order_by('-id')
    for obj in queryset:
        row_num += 1
        core_events = ""
        if obj.marketing:
            core_events = core_events + "MKT" + " "
        if obj.finance:
            core_events = core_events + "FIN" + " "
        if obj.public_relations:
            core_events = core_events + "PR" + " "
        if obj.human_resources:
            core_events = core_events + "HR" + " "
        if obj.ent_dev:
            core_events = core_events + "ED" + " "
        if obj.best_manager:
            core_events = core_events + "BM" + " "
        if obj.corp_strg:
            core_events = core_events + "CS" + " "
        if obj.quiz:
            core_events = core_events + "QU" + " "
        if obj.team:
            core_events = core_events + "TEAM" + " "
        if event.inter_type:
            row = [
                obj.rcode,
                obj.student.name,
                obj.student.email,
                obj.student.phoneno,
                obj.student.student_detail.degree,
                obj.student.student_detail.year,
                obj.student.student_detail.college,
                core_events,
            ]
        else:
            row = [
                obj.rcode,
                obj.student.name,
                obj.student.email,
                obj.student.phoneno,
                obj.student.student_detail.degree,
                obj.student.student_detail.year,
                obj.student.student_detail.section,
                core_events,
            ]
        for col_num in range(0, len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    wb.save(response)
    return response


def export_roundScore(request,id=None):
    import openpyxl
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=RoundScores.xlsx'
    round = rounds.objects.get(id=id)
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = round.title
    row_num = 0
    rcode = ""
    columns = [
        (u"ID", 80),
        (u"Name", 20),
        (u"Email", 20),
        (u"Phone", 20),
        (u"Course", 20),
        (u"Year", 20),
        (u"Section", 20),
        (u"Score", 20),
    ]
    for col_num in range(0, len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
    queryset = round_scores.objects.filter(round=round).order_by('rcode')
    for obj in queryset:
        if rcode != obj.rcode:
            rcode = obj.rcode
            row_num += 1
            row = [
                obj.rcode,
                obj.student.name,
                obj.student.email,
                obj.student.phoneno,
                obj.student.student_detail.degree,
                obj.student.student_detail.year,
                obj.student.student_detail.section,
                obj.total,
            ]
            for col_num in range(0, len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
    wb.save(response)
    return response


def event_feed(request):
    current_user = request.user
    club = clubs.objects.get(club_email=current_user.email)
    try:
        event = events.objects.get(email=club,current=True)
    except:
        event = None
    try:
        all_events = events.objects.filter(current=True, registration=True,inter_type=True)
        registered,created = event_registered.objects.get_or_create(current_user=current_user)
        registered = registered.registered_to.all()
        for event in all_events:
            event.flag = True
            if event in registered:
                event.flag = False
        #For when we need to register to club first(works, KEEP THIS CODE)
        # register = register_table.objects.get(current_user=current_user)
        # clubs = register.registered_to.all()
        # for x in clubs:
        #     y = events.objects.filter(email = x, current=True, live=True, registration=True)
        #     if all_events is None:
        #         all_events = y
        #     else:
        #         all_events = all_events | y
    except:
        all_events = None
    context = { "user": current_user, "events": all_events, "event":event}
    return render(request, 'club_dash/event_feed.html', context)


def student_list(request):
    user =  request.user
    club = clubs.objects.get(club_email=user.email)
    college = club.email
    try:
        event = events.objects.get(email=club,current=True)
    except:
        event = None
    students = student_detail.objects.filter(college=college.college_name,email__club=False,email__judge=False,email__college=False)
    friends,created = follow_table.objects.get_or_create(current_user=user)
    friends = friends.connected_to.all()
    for stu in students:
        stu.flag = False
        for friend in friends:
            if stu.email == friend:
                stu.flag = True
    context = {'user':user,'event':event,'students':students}
    return render(request,'club_dash/student_list.html',context)


def event_register(request, id = None):
    current_user = request.user
    club = clubs.objects.get(club_email=current_user.email)
    event = events.objects.get(id=id)
    try:
        network = follow_table.objects.get(current_user=current_user)
        network = network.connected_to.all()
    except:
        network = None
    if request.POST:
        event_registered.register(current_user, event)
        rcode = request.POST.get("prefix","")
        for x in range(0,event.team_size1):
            member = request.POST.get("mkt"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.marketing = True
            regdetails.regmarketing = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event,club=club, student=member, type=1)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size2):
            member = request.POST.get("fin"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.finance = True
            regdetails.regfinance = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=2)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size3):
            member = request.POST.get("pr"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.public_relations = True
            regdetails.regpublic_relations = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=3)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size4):
            member = request.POST.get("hr"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.human_resources = True
            regdetails.reghuman_resources = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=4)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size5):
            member = request.POST.get("ed"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.ent_dev = True
            regdetails.regent_dev = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=5)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size6):
            member = request.POST.get("bm"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.best_manager = True
            regdetails.regbest_manager = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=6)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size7):
            member = request.POST.get("cs"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.corp_strg = True
            regdetails.regcorp_strg = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=7)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size8):
            member = request.POST.get("qu"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.quiz = True
            regdetails.regquiz = True
            # regdetails.team = True
            # regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=8)
            team.tcode = rcode
            team.save()
        for x in range(0,event.team_size9):
            member = request.POST.get("te"+str(x))
            member = student.objects.get(email=member)
            event_registered.register(member, event)
            regdetails , created = event_registered_details.objects.get_or_create(event=event, student=member)
            regdetails.team = True
            regdetails.regteam = True
            regdetails.rcode = rcode
            regdetails.save()
            team, created = teams.objects.get_or_create(event=event, club=club, student=member, type=9)
            team.tcode = rcode
            team.save()
        return HttpResponseRedirect("/user/club_dashboard/eventFeed")
    context = {'event':event, "network":network,'team1':range(0,event.team_size1),'team2':range(0,event.team_size2),
               'team3': range(0, event.team_size3),'team4':range(0,event.team_size4),'team5':range(0,event.team_size5),
               'team6': range(0, event.team_size6),'team7':range(0,event.team_size7),'team8':range(0,event.team_size8),
               'team9': range(0, event.team_size9),}
    return render(request,'club_dash/register_details.html',context)


def network(request):
    user = request.user
    club = clubs.objects.get(club_email=user.email)
    college = club.email
    try:
        event = events.objects.get(email=club, current=True)
    except:
        event = None
    try:
        networks,created = follow_table.objects.get_or_create(current_user=user)
        networks = networks.connected_to.all()
    except:
        networks = None
    context = {'user': user, 'event': event, 'students': networks}
    return render(request, 'club_dash/network.html', context)

